from asyncio import Task
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.hashers import make_password, check_password
from .models import  User, OTP , Record
from django.conf import settings
from django.contrib.auth import login as auth_login
from .forms import JobApplicationFormm, UploadFileForm
import logging
import pandas as pd
import random
from django.core.mail import send_mail
from django.shortcuts import get_object_or_404, redirect
from .forms import UserSearchForm
from django.db.models import Q
from .forms import UploadFileForm
from .forms import OTPVerificationForm, PasswordResetForm
from .forms import EditProfileForm
from django.contrib.auth import logout
from django.contrib.auth.views import LogoutView
from django.contrib.auth.models import User
from django.shortcuts import redirect
from .forms import JobForm
from django.core.files.storage import FileSystemStorage


# Set up logging
logger = logging.getLogger(__name__)
def home(request):
    return render(request, "home.html")
def LoginPage(request): 
    # Render the login page
    return render(request, "login.html")

def verify_emailPage(request):
    # Render the forgot password page to verify email
    return render(request, "verify_email.html")

def index(request):
    # Render the index page
    return render(request, "index.html")

def signup(request):
    if request.method == "POST": 
        # Validate the form method and fetch the inputs
        firstname = request.POST.get('firstname', '').strip()
        lastname = request.POST.get('lastname', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        password = request.POST.get('password', '').strip()
        cPassword = request.POST.get('cpassword', '').strip()

        # Validate the inputs
        if not firstname or not lastname:
            messages.warning(request, "Enter valid names! Please check them.")
            return redirect("signuppage")
        if not email:
            messages.warning(request, "Email should not be empty!")
            return redirect("signuppage")
        if not phone or len(phone) != 10:
            messages.warning(request, "Phone number should be of 10 digits only!")
            return redirect("signuppage")
        if not password or not cPassword:
            messages.warning(request, "Password and Confirm Password are required.")
            return redirect("signuppage")
        if len(password) < 6 or len(cPassword) < 6:
            messages.warning(request, "Password and Confirm Password must require min. 6 digits.")
            return redirect("signuppage")
        if password != cPassword:
            messages.warning(request, "Password and Confirm Password should be the same.")
            return redirect("signuppage")

        # Check whether the email is already in use
        user_exists = User.objects.filter(email=email).exists()
        if user_exists:
            messages.error(request, "Email is already taken! Try with another!")
            return redirect("signuppage")
        else:
            # Create the user
            new_user = User.objects.create(
                firstname=firstname,
                lastname=lastname,
                email=email,
                phone=phone,
                password=make_password(password)  # encrypted password
            )
            if new_user:
                messages.success(request, "Signup successful!")
                return redirect("loginPage")
            else:
                messages.error(request, "Error in creating new user! Try again.")
                return redirect("signuppage")
    else:
        messages.error(request, "Bad request method!")
        return redirect("signuppage")

def signupPage(request):
    # Render the signup page
    return render(request, "signuppage.html")

def main(request):
    # Render the signup page
    return render(request, "main.html")

def about(request):
    return render(request, "about.html")
def jobListing(request): 
    # Render the login page
    return render(request, "job-list.html")

def exam(request):
    # Render the signup page
    return render(request, "exam.html")
from django.contrib.auth import authenticate, login as auth_login

def login_auth(request):
    if request.method == "POST":
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()

        user = authenticate(request, username=email, password=password)
        if user is not None:
            auth_login(request, user)
            messages.success(request, "Login successful")
            return redirect('mainpage')
        else:
            messages.error(request, "Invalid credentials")
            return redirect('loginPage')
    else:
        messages.error(request, "Invalid request method")
        return redirect('loginPage')

def show(request):
    users = User.objects.all()
    return render(request, "show.html", {'users': users})

def edit(request, id):
    user = get_object_or_404(User, id=id)
    return render(request, "edit.html", {'user': user})

def update(request, id):
    user = get_object_or_404(User, id=id)
    if request.method == "POST":
        user.email = request.POST.get('email', '').strip()
        user.password = make_password(request.POST.get('password', '').strip())
        user.firstname = request.POST.get('firstname', '').strip()
        user.lastname = request.POST.get('lastname', '').strip()
        user.phone = request.POST.get('phone', '').strip()
        user.save()
        return redirect("show")
    return render(request, 'edit.html', {'user': user})

def destroy(request, id):
    user = get_object_or_404(User, id=id)
    user.delete()
    return redirect("show")
from openpyxl import Workbook
from openpyxl.styles import Font
from django.http import HttpResponse

def generate_excel_report_task(request):
    tasks = User.objects.all()

    # Create a new workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Task Report'

    # Add header row
    header = ["firstname", "lastname", "email", "phone"]
    sheet.append(header)

    header_cells = sheet[1]
    for cell in header_cells:
        cell.font = Font(bold=True)
    # Add data rows
    for task in tasks:
        row = [
            task.firstname,
            task.lastname,
            task.email,
            task.phone
        ]
        sheet.append(row)

    # Create the HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="task_report.xlsx"'

    # Save the workbook to the response
    workbook.save(response)
    return response
# Generate a random OTP
#Excel for applicants

def verify_email(request):
    if request.method == "POST":
        email = request.POST.get('email', '').strip()
        if not email:
            messages.error(request, "Please enter a valid email address.")
            return redirect('verify_email')

        otp_code = random.randint(100000, 999999)
        OTP.objects.create(email=email, otp=otp_code)

        send_mail(
            'Your OTP Code',
            f'Your OTP code is {otp_code}',
            'from@example.com',
            [email],
            fail_silently=False,
        )

        messages.success(request, "OTP sent to your email.")
        return redirect('verify_otp')
    
    return render(request, 'verify_email.html')

def verify_otp(request):
    if request.method == "POST":
        form = OTPVerificationForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            otp_code = form.cleaned_data['otp_code']
            
            try:
                otp_record = OTP.objects.get(email=email, otp=otp_code)
                otp_record.delete()  # Optionally delete OTP record after successful verification
                messages.success(request, "OTP verified successfully.")
                return redirect(f'/reset_password/?email={email}')
            except OTP.DoesNotExist:
                messages.error(request, "Invalid OTP or email.")
                return redirect('verify_otp')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = OTPVerificationForm()

    return render(request, 'verify_otp.html', {'form': form})



from django.shortcuts import redirect
from django.contrib.auth import get_user_model

User = get_user_model()

def reset_password(request):
    if request.method == "POST":
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            new_password = form.cleaned_data['new_password']

            try:
                user = User.objects.get(email=email)
                user.password = make_password(new_password)
                user.save()
                messages.success(request, "Password reset successfully. Please log in with your new password.")
                return redirect('loginPage')  # Ensure this URL name is correct
            except User.DoesNotExist:
                messages.error(request, "User not found.")
                return redirect('reset_password')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        email = request.GET.get('email')
        if not email:
            messages.error(request, "Invalid request.")
            return redirect('loginPage')  # Ensure this URL name is correct

        form = PasswordResetForm(initial={'email': email})

    return render(request, 'reset_password.html', {'form': form})

#Search
def search_users(request):
    form = UserSearchForm()
    results = []
    
    if request.method == 'GET':
        form = UserSearchForm(request.GET)
        if form.is_valid():
            query = form.cleaned_data['query']
            results = User.objects.filter(
                Q(firstname__icontains=query) |
                Q(lastname__icontains=query) |
                Q(email__icontains=query) |
                Q(phone__icontains=query)
            )
    
    return render(request, 'search_results.html', {'form': form, 'results': results})

#Upload
def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            df = pd.read_excel(file)
            for index, row in df.iterrows():
                Record.objects.create(
                    email=row['email'],
                    #age=row['age'],
                    name=row['name']
                )
            return redirect('upload')
    else:
        form = UploadFileForm()
    return render(request, 'upload.html', {'form': form})

def download_file(request):
    records = Record.objects.all().values()
    df = pd.DataFrame(records)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=records.xlsx'
    df.to_excel(response, index=False)
    return response

def edit_profile(request):
    if request.method == 'POST':
        form = EditProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your profile has been updated successfully.')
            return redirect('mainpage')  # Redirect to the main page or wherever you want
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = EditProfileForm(instance=request.user)

    return render(request, 'edit_profile.html', {'form': form})

def user_logout(request):
    logout(request)
    return redirect('loginPage')

#Admin
from django.contrib.auth.views import LoginView
from django.urls import path

class CustomLoginView(LoginView):
    def get_success_url(self):
        if self.request.user.is_staff:
            return '/admin/'  # Redirect staff to admin panel
        else:
            return '/signuppage/' 
        
from django.contrib.auth.views import LogoutView
from django.urls import reverse_lazy

class CustomAdminLogoutView(LogoutView):
    def get_next_page(self):
        # Redirect to the admin login page after admin logout
        if self.request.path.startswith('/admin/'):
            return '/admin/login/'  # Admin login page URL
        return super().get_next_page()
    
#Job List
from .forms import JobApplicationForm
from .models import JobApplication

from django.contrib import messages  # Import Django messages

def apply_job(request):
    if request.method == 'POST':
        form = JobApplicationForm(request.POST, request.FILES)
        if form.is_valid():
            # Save form data to the database
            JobApplication.objects.create(
                firstname=form.cleaned_data['firstname'],
                lastname=form.cleaned_data['lastname'],
                email=form.cleaned_data['email'],
                highest_qualification=form.cleaned_data['highest_qualification'],
                course_specialization=form.cleaned_data['course_specialization'],
                experience=form.cleaned_data['experience'],
                company=form.cleaned_data.get('company', ''),
                cv=form.cleaned_data['cv']
            )
            # Add success message
            messages.success(request, 'Details uploaded successfully!')
            # Redirect back to the same page
            #return redirect(request.path)
            return redirect('mainpage')
    else:
        if request.user.is_authenticated:
            form = JobApplicationForm(initial={
                'firstname': request.user.firstname,
                'lastname': request.user.lastname,
                'email': request.user.email
            })
        else:
            form = JobApplicationForm()

    return render(request, 'job-detail.html', {'form': form})

#View the application
from django.contrib.auth.decorators import login_required

@login_required
def view_applications(request):
    applications = JobApplication.objects.all()
    return render(request, 'admin/view_applications.html', {'applications': applications})

from .models import Job

# View for user to see all available jobs
def post_job(request):
    if request.method == 'POST':
        form = JobForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('job_list')  # Redirect to the job listing page
    else:
        form = JobForm()
    return render(request, 'jobs/post_job.html', {'form': form})

# View for the detailed job description (for users)
def job_list(request):
    query = request.GET.get('q')  # Get the search query from the request
    if query:
        jobs = Job.objects.filter(
            Q(title__icontains=query) | 
            Q(company_name__icontains=query) | 
            Q(location__icontains=query) |
            Q(job_type__icontains=query)
        )
    else:
        jobs = Job.objects.all()

    # Admin template and user template logic remains the same
    if request.user.is_superuser:
        template_name = 'jobs/admin_job_listings.html'
        print("User is an admin")
    else:
        template_name = 'jobs/user_job_listings.html'
        print("User is a regular user")

    return render(request, template_name, {'jobs': jobs})

    
def job_detail(request, job_id):
    job = get_object_or_404(Job, id=job_id)
    return render(request, 'jobs/job_detail.html', {'job': job})

# views.py

def applyjob(request, job_id):
    job = get_object_or_404(Job, id=job_id)  # Fetch job details
    jobs = Job.objects.all()
    
    if request.method == 'POST':
        form = JobApplicationFormm(request.POST, request.FILES)
        if form.is_valid():
            application = form.save(commit=False)
            application.job = job  # Assign the job to the application
            application.save()
            return render(request, 'jobs/job_list.html', {'jobs': jobs})  # Redirect after submission
    else:
        # Pre-fill the form with job details
        form = JobApplicationFormm(initial={
            'company': job.company_name,
            'location': job.location,
            'job_type': job.job_type,
            'job_description': job.description,
        })

    return render(request, 'jobs/applyjob.html', {'form': form, 'job': job})


#Manage Jobs
from .forms import JobForm  # Import your JobForm

def edit_job(request, job_id):
    job = get_object_or_404(Job, id=job_id)

    if request.method == 'POST':
        form = JobForm(request.POST, instance=job)
        if form.is_valid():
            form.save()
            return redirect('job_list')  # Redirect to the job list after saving
    else:
        form = JobForm(instance=job)

    return render(request, 'jobs/edit_job.html', {'form': form, 'job': job})

def delete_job(request, job_id):
    job = get_object_or_404(Job, id=job_id)
    
    if request.method == 'POST':
        job.delete()
        return redirect('job_list')  # Redirect to the job list after deleting

    return render(request, 'jobs/delete_job.html', {'job': job})

#Admin Numbers


from django.utils import timezone
from django.db.models import Count

def admin_dashboard(request):
    # Fetch the total number of jobs
    total_jobs = Job.objects.count()
    
    # Fetch the total number of job applications
    total_applications = JobApplication.objects.count()

    # Fetch the number of applicants for each month (assuming 'date_applied' exists in JobApplication)
    applicants_over_time = JobApplication.objects.filter(
        date_applied__year=timezone.now().year
    ).annotate(month=ExtractMonth('date_applied')).values('month').annotate(total=Count('id')).order_by('month') # type: ignore

    # Fetch the number of jobs available for each month
    jobs_over_time = Job.objects.filter(
        created_at__year=timezone.now().year
    ).annotate(month=ExtractMonth('created_at')).values('month').annotate(total=Count('id')).order_by('month') # type: ignore

    context = {
        'total_jobs': total_jobs,
        'total_applications': total_applications,
        'applicants_over_time': list(applicants_over_time),
        'jobs_over_time': list(jobs_over_time),
    }

    return render(request, 'admin/index.html', context)

#Graph
from django.db.models.functions import ExtractMonth
@login_required
def admin_dashboard(request):
    total_jobs = Job.objects.count()
    total_applications = JobApplication.objects.count()

    applicants_over_time = JobApplication.objects.filter(
        date_applied__year=timezone.now().year
    ).annotate(month=ExtractMonth('date_applied')).values('month').annotate(total=Count('id')).order_by('month')

    jobs_over_time = Job.objects.filter(
        created_at__year=timezone.now().year
    ).annotate(month=ExtractMonth('created_at')).values('month').annotate(total=Count('id')).order_by('month')

    context = {
        'total_jobs': total_jobs,
        'total_applications': total_applications,
        'applicants_over_time': list(applicants_over_time),
        'jobs_over_time': list(jobs_over_time),
    }

    return render(request, 'admin/index.html', context)


# views.py
from django.shortcuts import render

# views.py
from django.shortcuts import render
from .models import ApplyJobApplication  # Import your model

def job_applications_view(request):
    # Fetch all job applications from the database
    job_applications = ApplyJobApplication.objects.all()

    # Pass the data to the template context
    return render(request, 'admin/job_applications.html', {'applications': job_applications})

import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from .models import ApplyJobApplication
from django.http import HttpResponse

# your_app_name/views.py
from django.http import HttpResponse
from .models import ApplyJobApplication
import openpyxl
from openpyxl.styles import Font

def excel_job_applicants(request):
    print("excel_job_applicants called")  # Debugging line to check function execution
    try:
        # Fetch all job applications from the database
        job_applications = ApplyJobApplication.objects.all()
        if not job_applications:
            print("No job applications found")  # Debugging line
            return HttpResponse("No job applications found.", status=404)

        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Job Applications'

        # Define the header row
        headers = ['First Name', 'Last Name', 'Email', 'Company', 'Location', 'Job Type', 'Job Description', 'CV']
        header_font = Font(bold=True)

        # Write headers to the first row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font

        # Write data rows
        for row_num, application in enumerate(job_applications, 2):
            ws.cell(row=row_num, column=1, value=application.firstname)
            ws.cell(row=row_num, column=2, value=application.lastname)
            ws.cell(row=row_num, column=3, value=application.email)
            ws.cell(row=row_num, column=4, value=application.company)
            ws.cell(row=row_num, column=5, value=application.location)
            ws.cell(row=row_num, column=6, value=application.job_type)
            ws.cell(row=row_num, column=7, value=application.job_description)
            ws.cell(row=row_num, column=8, value=application.cv.url)

        # Set the response to return as an Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=job_applications.xlsx'

        # Save the workbook to the response
        wb.save(response)

        return response
    except Exception as e:
        print(f"An error occurred: {e}")  # Print any errors
        return HttpResponse("An error occurred while generating the Excel file.", status=500)

def test_excel_export(request):
    print("test_excel_export called")  # Debugging line
    return excel_job_applicants(request)

# your_app_name/views.py
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from .models import Job  # Import your Job model

def excel_job_list(request):
    # Fetch all job listings from the database
    jobs = Job.objects.all()

    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Job Listings'

    # Define the header row
    headers = ['Title', 'Company Name', 'Location', 'Job Type', 'Description', 'Who Can Apply', 'Posted At', 'Last Date to Apply']
    header_font = Font(bold=True)  # Make the header bold

    # Write headers to the first row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font  # Apply bold font to the headers

    # Write data rows
    for row_num, job in enumerate(jobs, 2):
        ws.cell(row=row_num, column=1, value=job.title)
        ws.cell(row=row_num, column=2, value=job.company_name)
        ws.cell(row=row_num, column=3, value=job.location)
        ws.cell(row=row_num, column=4, value=job.get_job_type_display())  # Get the display name for job type
        ws.cell(row=row_num, column=5, value=job.description)
        ws.cell(row=row_num, column=6, value=job.who_can_apply)
        ws.cell(row=row_num, column=7, value=job.posted_at.strftime('%Y-%m-%d %H:%M:%S'))  # Format date
        ws.cell(row=row_num, column=8, value=job.last_date_to_apply)  # Assuming last_date_to_apply is a date

    # Set the response to return as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=job_listings.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response

def test_excel_job(request):
    print("test_excel_export called")  # Debugging line
    return excel_job_list(request)


#Mails
from django.core.mail import send_mail
from django.shortcuts import redirect
from django.contrib import messages
from .models import JobApplication  # Adjust according to your model name

def send_email(request, applicant_id):
    try:
        applicant = JobApplication.objects.get(id=applicant_id)
        subject = "Job Application Status"
        message = f"Dear {applicant.firstname} {applicant.lastname},\n\nThank you for applying to {applicant.company}. We are pleased to inform you that your application has been approved. Welcome aboard!\n\nBest regards,\n[Your Company]"
        recipient_list = [applicant.email]
        send_mail(subject, message, 'your_email@example.com', recipient_list)
        messages.success(request, f"Email sent to {applicant.firstname} {applicant.lastname}.")
    except JobApplication.DoesNotExist:
        messages.error(request, "Applicant not found.")
    except Exception as e:
        messages.error(request, f"An error occurred: {e}")
    return redirect('job_applications')  # Adjust to your actual view name
